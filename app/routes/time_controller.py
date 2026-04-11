from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from app.db.database import get_db
from app.models.dto import User, TimeEntry
from app.auth import get_current_user
import datetime
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook

router = APIRouter()

def entry_to_dict(entry):
    return {
        "id": entry.id,
        "user_id": entry.user_id,
        "clock_in": entry.clock_in.isoformat() + 'Z' if entry.clock_in else None,
        "clock_out": entry.clock_out.isoformat() + 'Z' if entry.clock_out else None,
    }

@router.post("/clock-in")
async def clock_in(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    existing = db.query(TimeEntry).filter(TimeEntry.user_id == current_user.id, TimeEntry.clock_out == None).first()
    if existing:
        raise HTTPException(status_code=400, detail="Already clocked in")
    entry = TimeEntry(user_id=current_user.id)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"message": "Clocked in", "entry": entry_to_dict(entry)}

@router.post("/clock-out")
async def clock_out(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entry = db.query(TimeEntry).filter(TimeEntry.user_id == current_user.id, TimeEntry.clock_out == None).first()
    if not entry:
        raise HTTPException(status_code=400, detail="Not clocked in")
    entry.clock_out = datetime.datetime.utcnow()
    db.commit()
    db.refresh(entry)
    return {"message": "Clocked out", "entry": entry_to_dict(entry)}

@router.get("/my-entries")
async def get_my_entries(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    entries = db.query(TimeEntry).filter(TimeEntry.user_id == current_user.id).order_by(TimeEntry.clock_in).all()
    return [entry_to_dict(e) for e in entries]

@router.get("/all-entries")
async def get_all_entries(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="Not authorized")
    entries = db.query(TimeEntry).order_by(TimeEntry.clock_in).all()
    return [entry_to_dict(e) for e in entries]

@router.get("/export-entries")
async def export_entries(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="Not authorized")
    entries = db.query(TimeEntry).join(User).order_by(TimeEntry.clock_in).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"
    
    # Headers
    ws['A1'] = 'User ID'
    ws['B1'] = 'User Name'
    ws['C1'] = 'User Email'
    ws['D1'] = 'Clock In'
    ws['E1'] = 'Clock Out'
    ws['F1'] = 'Duration (hours)'
    
    for i, entry in enumerate(entries, start=2):
        duration = None
        if entry.clock_out:
            duration = (entry.clock_out - entry.clock_in).total_seconds() / 3600
        ws[f'A{i}'] = entry.user_id
        ws[f'B{i}'] = entry.user.full_name or 'N/A'
        ws[f'C{i}'] = entry.user.email
        ws[f'D{i}'] = entry.clock_in.strftime('%Y-%m-%d %H:%M:%S') if entry.clock_in else ''
        ws[f'E{i}'] = entry.clock_out.strftime('%Y-%m-%d %H:%M:%S') if entry.clock_out else ''
        ws[f'F{i}'] = round(duration, 2) if duration else ''
    
    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    return StreamingResponse(
        bio,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=time_entries.xlsx"}
    )